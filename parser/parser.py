"""
Parser module for Excel schedule files.
Supports badly formatted XLSX files with multiple fallback strategies.
"""

import logging
import os
from pathlib import Path
from typing import List, Dict, Any, Optional
import re
from datetime import datetime

logger = logging.getLogger(__name__)

# Try importing parsers in order of preference
try:
    from python_calamine import CalamineWorkbook
    CALAMINE_AVAILABLE = True
    logger.info("python-calamine is available")
except ImportError:
    CALAMINE_AVAILABLE = False
    logger.warning("python-calamine not available, will use openpyxl")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
    logger.info("openpyxl is available")
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
    logger.info("pandas is available")
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("pandas not available")


class ScheduleParser:
    """Parse college schedule from Excel files with multiple fallback strategies."""
    
    def __init__(self):
        """Initialize the parser."""
        self.group_pattern = re.compile(r'([А-Я]{1,3}[0-9]?-\d{2})')
        self.time_pattern = re.compile(r'(\d{1,2}:\d{2}[-–]\d{1,2}:\d{2})')
        self.date_pattern = re.compile(r'(\d{1,2}\.\d{1,2}\.\d{4})')
        self.weekdays = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье']
    
    def parse_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Parse an Excel file and extract schedule information.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of schedule entries
        """
        logger.info(f"Parsing file: {file_path}")
        
        # Try calamine first (fastest and most robust)
        if CALAMINE_AVAILABLE:
            try:
                return self._parse_with_calamine(file_path)
            except Exception as e:
                logger.warning(f"Calamine parsing failed: {e}, trying openpyxl")
        
        # Fallback to openpyxl
        if OPENPYXL_AVAILABLE:
            try:
                return self._parse_with_openpyxl(file_path)
            except Exception as e:
                logger.warning(f"Openpyxl parsing failed: {e}, trying pandas")
        
        # Fallback to pandas
        if PANDAS_AVAILABLE:
            try:
                return self._parse_with_pandas(file_path)
            except Exception as e:
                logger.error(f"Pandas parsing failed: {e}")
        
        logger.error(f"All parsing methods failed for {file_path}")
        return []
    
    def _parse_with_calamine(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse using python-calamine (fastest)."""
        workbook = CalamineWorkbook.from_path(file_path)
        all_entries = []
        
        for sheet_name in workbook.sheet_names:
            logger.info(f"Processing sheet: {sheet_name}")
            data = workbook.get_sheet_by_name(sheet_name).to_python()
            entries = self._extract_schedule_data(data, sheet_name, Path(file_path).name)
            all_entries.extend(entries)
        
        logger.info(f"Extracted {len(all_entries)} entries using calamine")
        return all_entries
    
    def _parse_with_openpyxl(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse using openpyxl."""
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        all_entries = []
        
        for sheet_name in workbook.sheetnames:
            logger.info(f"Processing sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            
            # Convert to list of lists
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            
            entries = self._extract_schedule_data(data, sheet_name, Path(file_path).name)
            all_entries.extend(entries)
        
        logger.info(f"Extracted {len(all_entries)} entries using openpyxl")
        return all_entries
    
    def _parse_with_pandas(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse using pandas."""
        all_entries = []
        
        # Get all sheet names
        excel_file = pd.ExcelFile(file_path)
        
        for sheet_name in excel_file.sheet_names:
            logger.info(f"Processing sheet: {sheet_name}")
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # Convert to list of lists
            data = df.values.tolist()
            
            entries = self._extract_schedule_data(data, sheet_name, Path(file_path).name)
            all_entries.extend(entries)
        
        logger.info(f"Extracted {len(all_entries)} entries using pandas")
        return all_entries
    
    def _extract_schedule_data(self, data: List[List], sheet_name: str, 
                               file_name: str) -> List[Dict[str, Any]]:
        """
        Extract schedule information from raw data.
        
        Args:
            data: Raw data as list of lists
            sheet_name: Name of the sheet
            file_name: Name of the file
            
        Returns:
            List of schedule entries
        """
        entries = []
        
        # Find header row with group names
        header_row_idx = self._find_header_row(data)
        if header_row_idx is None:
            logger.warning(f"Could not find header row in {sheet_name}")
            return entries
        
        # Extract group columns mapping
        groups_mapping = self._extract_groups_mapping(data[header_row_idx])
        if not groups_mapping:
            logger.warning(f"Could not find groups in {sheet_name}")
            return entries
        
        logger.info(f"Found groups: {list(groups_mapping.keys())}")
        
        # Parse schedule rows
        current_day = None
        current_date = None
        
        for row_idx in range(header_row_idx + 1, len(data)):
            row = data[row_idx]
            if not row or all(self._is_empty(cell) for cell in row):
                continue
            
            # Check for day of week
            day = self._extract_day_of_week(row)
            if day:
                current_day = day
                current_date = self._extract_date_from_row(row)
                logger.debug(f"Found day: {current_day}, date: {current_date}")
                continue
            
            # Check for time slot
            time_slot = self._extract_time_slot(row)
            if not time_slot or not current_day:
                continue
            
            # Extract lessons for each group
            for group_name, col_indices in groups_mapping.items():
                subject_col, room_col = col_indices
                
                subject = self._get_cell_value(row, subject_col)
                room = self._get_cell_value(row, room_col)
                
                # Try to get teacher from next row
                teacher = None
                if row_idx + 1 < len(data):
                    next_row = data[row_idx + 1]
                    if not self._extract_time_slot(next_row) and not self._extract_day_of_week(next_row):
                        teacher = self._get_cell_value(next_row, subject_col)
                
                if subject and not self._is_empty(subject):
                    entry = {
                        'file_name': file_name,
                        'sheet_name': sheet_name,
                        'group_name': group_name,
                        'course': self._extract_course_from_name(group_name),
                        'day_of_week': current_day,
                        'date': current_date,
                        'time_slot': time_slot,
                        'subject': self._clean_text(subject),
                        'room': self._clean_text(room) if room else None,
                        'teacher': self._clean_text(teacher) if teacher else None
                    }
                    entries.append(entry)
        
        return entries
    
    def _find_header_row(self, data: List[List]) -> Optional[int]:
        """Find the row containing group names."""
        for idx, row in enumerate(data):
            if not row:
                continue
            
            # Look for typical header patterns
            row_text = ' '.join([str(cell) for cell in row if cell and not self._is_empty(cell)])
            
            # Check if this row has group names
            if self.group_pattern.search(row_text):
                # Make sure it also has "ауд" or similar markers
                if 'ауд' in row_text.lower() or 'группа' in row_text.lower():
                    return idx
        
        return None
    
    def _extract_groups_mapping(self, header_row: List) -> Dict[str, tuple]:
        """
        Extract mapping of group names to their column indices.
        
        Returns:
            Dict with group_name: (subject_col, room_col)
        """
        groups = {}
        i = 0
        
        while i < len(header_row):
            cell = header_row[i]
            if cell and not self._is_empty(cell):
                cell_str = str(cell)
                match = self.group_pattern.search(cell_str)
                
                if match:
                    group_name = match.group(1)
                    subject_col = i
                    room_col = i + 1 if i + 1 < len(header_row) else i
                    
                    groups[group_name] = (subject_col, room_col)
                    logger.debug(f"Found group {group_name} at columns {subject_col}, {room_col}")
                    i += 2  # Skip room column
                    continue
            
            i += 1
        
        return groups
    
    def _extract_day_of_week(self, row: List) -> Optional[str]:
        """Extract day of week from row."""
        for cell in row[:6]:  # Days usually in first few columns
            if cell and not self._is_empty(cell):
                cell_str = str(cell).lower().strip()
                for day in self.weekdays:
                    if day in cell_str:
                        return day
        return None
    
    def _extract_date_from_row(self, row: List) -> Optional[str]:
        """Extract date from row and convert to YYYY-MM-DD format."""
        for cell in row:
            if cell and not self._is_empty(cell):
                cell_str = str(cell)
                
                # Check for datetime object
                if hasattr(cell, 'strftime'):
                    try:
                        return cell.strftime('%Y-%m-%d')
                    except:
                        pass
                
                # Check for date string
                match = self.date_pattern.search(cell_str)
                if match:
                    date_str = match.group(1)
                    try:
                        # Parse DD.MM.YYYY format
                        dt = datetime.strptime(date_str, '%d.%m.%Y')
                        return dt.strftime('%Y-%m-%d')
                    except:
                        pass
        
        return None
    
    def _extract_time_slot(self, row: List) -> Optional[str]:
        """Extract time slot from row."""
        for cell in row[:6]:  # Time usually in first few columns
            if cell and not self._is_empty(cell):
                cell_str = str(cell)
                if self.time_pattern.search(cell_str):
                    return self._clean_text(cell_str)
        return None
    
    def _extract_course_from_name(self, group_name: str) -> Optional[int]:
        """Extract course number from group name (e.g., БУ1-24 -> 1)."""
        # Look for year pattern like -24, -25
        year_match = re.search(r'-(\d{2})', group_name)
        if year_match:
            year = int(year_match.group(1))
            current_year = datetime.now().year % 100
            # Calculate course based on year difference
            course = current_year - year + 1
            if 1 <= course <= 4:
                return course
        
        # Try to find course number in the name
        course_match = re.search(r'(\d+)', group_name)
        if course_match:
            course = int(course_match.group(1))
            if 1 <= course <= 4:
                return course
        
        return None
    
    def _get_cell_value(self, row: List, col_idx: int):
        """Safely get cell value from row."""
        if col_idx < len(row):
            return row[col_idx]
        return None
    
    def _is_empty(self, value) -> bool:
        """Check if cell value is empty."""
        if value is None:
            return True
        
        value_str = str(value).strip().lower()
        
        # Check for common empty values
        if value_str in ['', 'nan', 'none', 'nat', 'null']:
            return True
        
        return False
    
    def _clean_text(self, text) -> str:
        """Clean and normalize text."""
        if text is None:
            return ""
        
        text_str = str(text).strip()
        
        # Remove extra whitespace
        text_str = re.sub(r'\s+', ' ', text_str)
        
        # Remove common artifacts
        text_str = text_str.replace('nan', '').replace('NaT', '')
        
        return text_str.strip()
    
    def find_schedule_files(self, folder_path: str, pattern: str = "*.xlsx") -> List[str]:
        """
        Find all Excel schedule files in a folder.
        
        Args:
            folder_path: Path to search for files
            pattern: File pattern to match
            
        Returns:
            List of file paths
        """
        folder = Path(folder_path)
        files = list(folder.glob(pattern))
        
        # Filter out temporary files
        files = [f for f in files if not f.name.startswith('~') and not f.name.startswith('.')]
        
        logger.info(f"Found {len(files)} schedule files in {folder_path}")
        return [str(f) for f in files]
